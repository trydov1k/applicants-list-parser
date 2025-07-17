import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

def take_settings():  # Возвращает текущие настройки
    standart_settings = {
        "folder_path": "None",
        "universities": [],
        "abit_id": 0}
    try:
        with open('settings.json', 'r') as json_file:
            settings = json.load(json_file)
        return settings
    except FileNotFoundError:
        with open('settings.json', 'w') as json_file:
            json.dump(standart_settings, json_file)
            return standart_settings

def change_settings(setting_name, new_value):  # Меняет текущие настройки, получает имя настройки и новое значение этой настройки
    settings = take_settings()
    settings[setting_name] = new_value
    with open('settings.json', 'w') as json_file:
        json.dump(settings, json_file)

def take_abit_id():  # Возвращает id абитуриента
    if take_settings()["abit_id"] == 0:
        try:
            ans = int(input('Хотите убрать участников, которые находятся в списке после вас? (1/0)\n'))
            if not ans:
                abit_id = -1
            else:
                try:       
                    abit_id = int(input('Введите свой ID участника\n'))
                except:
                    print('Что-то пошло не так, попробуйте ещё раз')
                    main()
        except:
            print('Что-то пошло не так, попробуйте ещё раз')
            main()
    else:
        abit_id = take_settings()["abit_id"]
    change_settings("abit_id", abit_id)
    return abit_id

def take_tables_paths():  # Возвращает пути до файлов (таблиц), которые нужно обработать
    work_folder_path = os.path.dirname(os.path.abspath(__file__))
    tables_paths = []
    if take_settings()['folder_path'] == 'None':
        try:
            ans = int(input('Хотите поменять папку с таблицами? (1/0)\n'))
            if ans:
                path_to_tables = input('Введите путь до папки, из которой будут взяты таблицы со списками поступающих'+'\n'+r'Пример: C:\Users\trydov1k\OneDrive\Desktop\Списки поступающих'+'\n')
                for root, dirs, files in os.walk(path_to_tables):
                    for file in files:
                        if file.endswith('.csv'):
                            try:
                                tables_paths.append(os.path.join(root, file))
                            except:
                                print('Вы ввели неверный путь, попробуйте ещё раз')
                                main()
                change_settings("folder_path", path_to_tables)
            else:
                pth = work_folder_path+'\\tables'
                change_settings("folder_path", pth)
                for root, dirs, files in os.walk(pth):
                    for file in files:
                        if file.endswith('.csv'):
                            tables_paths.append(os.path.join(root, file))
        except:
            print('Что-то пошло не так, попробуйте ещё раз')
            main()
    else:
        for root, dirs, files in os.walk(take_settings()['folder_path']):
                    for file in files:
                        if file.endswith('.csv'):
                            tables_paths.append(os.path.join(root, file))
    return tables_paths

def directions_to_universities():  # Меняет привязку специальностей к университету
    if len(take_settings()["universities"]) == 0:
        try:
            ans = int(input('Хотите распределять направления по университетам? (1/0)\n'))
            if not ans:
                dirrs = []
                try:
                    dirr_count = int(input('Введите количество специальностей (обрабатываемых файлов)\n'))
                    for nnn in range(dirr_count):
                        dirrs.append(input(f'{nnn+1}) '))
                except:
                    print('Что-то пошло не так, попробуйте ещё раз')
                    main()
                change_settings("universities", [["AllDirections", dirrs]])
            else:
                try:
                    count_universities = int(input('Введите количество университетов (от 1 до 5)\n'))
                    universities = []
                    result = []
                    print('Введите названия университетов, каждое на отдельной строчке')
                    for n in range(count_universities):
                        university = input(f'{n+1}) ')
                        universities.append(university)
                    for university in universities:
                        count_directions = int(input(f'Введите количество специальностей (количество обрабатываемых таблиц) для университета {university}\n'))
                        directions = []
                        print(f'Введите названия специальностей (таблиц) для университета {university}, каждое на отдельной строчке')
                        for nn in range(count_directions):
                            direction = input(f'{nn+1}) ')
                            directions.append(direction)
                        result.append([university, directions])
                    change_settings("universities", result)
                except:
                    print('Что-то пошло не так, попробуйте ещё раз')
                    main()
        except:
            print('Что-то пошло не так, попробуйте ещё раз')
            main()

def create_sample_books(diction, universities, columns_names):  # Создаёт образец(цы) для готовых файлов
    work_folder_path = os.path.dirname(os.path.abspath(__file__))
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    for book_name in universities:
        wb = Workbook()
        
        for sheet_name in diction[book_name]:
            wb.create_sheet(sheet_name)
            sheet = wb[sheet_name]
            for n in range(len(columns_names)):
                sheet[alphabet[n]+'1'] = columns_names[n]
                sheet.column_dimensions[alphabet[n]].width = 20
            
        wb.remove(wb["Sheet"])

        

        wb.save(work_folder_path + "\\sampletables\\" + book_name + '_sample.xlsx')

def fill_tables(tables, dictionary, abit_id):  # Заполняет таблицу(ы) данными
    work_folder_path = os.path.dirname(os.path.abspath(__file__))
    alphabet = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    for university in list(dictionary.keys()):
        wb = load_workbook(work_folder_path+'\\sampletables\\'+university+'_sample.xlsx')
        for table in tables:
            direction_name = table[0]
            columns_names = table[1]
            data = table[2:]
            if direction_name in dictionary[university]:
                a = 1
            else:
                continue
            
            sheet = wb[direction_name]
            
            dt = []
            for d in data:
                d = [int(d[0]), int(d[1]), d[2], int(d[3]), int(d[4]), int(d[5])]
                dt.append(d)

            for row in range(2, len(data)+2):
                for n in range(len(columns_names)):
                    column = alphabet[n]
                    if dt[row-2][3] == abit_id:
                        sheet[column+str(row)].font= Font(color="069A2E", bold=True)
                    sheet[column+str(row)] = dt[row-2][n]

        wb.save(work_folder_path+'\\readytables\\'+university+'_ready.xlsx')


def main():
    tables_paths = take_tables_paths()
    abit_id = take_abit_id()

    new_tables = []
    for path in tables_paths:
        table = open(path, encoding='utf-8')
        table_name = path.split("\\")[-1].split('.')[0]
        
        columns_names = [name.replace('\n', '').replace('\ufeff', '') for name in table.readline().replace('"', '').split(';')]
        rows = [row.replace('\n', '').replace('"', '').split(';') for row in table.readlines()]
        new_table = [table_name, [columns_names[0], columns_names[1], columns_names[6], columns_names[7], columns_names[5], columns_names[3]]]
        u = 0
        k = 0
        for row in rows:
            if row[4] == 'Без вступительных испытаний': row[4] = "400 100"
            if row[6] == 'Конкурсная группа исключена':
                continue
            sm = int(row[5])+sum(int(i) for i in row[4].split())
            new_table.append([row[0], row[1], row[6], row[7], row[5], sm])
            if int(row[7]) == abit_id:
                u = 1
            if u == 1:
                k += 1
            if k == 6:
                break
        new_tables.append(new_table)
       
    directions_to_universities()

    directions_to_universities_dict = {i[0]: i[1] for i in take_settings()["universities"]}
    
    all_universities_list = list(directions_to_universities_dict.keys())
    all_directions_list = list(directions_to_universities_dict.values())
    for i in [i[1] for i in take_settings()["universities"]]:
        for ii in i: all_directions_list.append(ii)
    
    

    create_sample_books(directions_to_universities_dict, all_universities_list, new_tables[0][1])
   
    fill_tables(new_tables, directions_to_universities_dict, abit_id)

if __name__ == '__main__':
    main()